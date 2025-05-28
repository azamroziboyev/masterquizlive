from aiogram import Bot, Dispatcher, types, F
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.filters import Command
from docx import Document
import asyncio
import os
import json
import logging
import time
import random
import signal
import concurrent.futures
from datetime import datetime
from functools import partial

from quiz_utils import convert_format, calculate_points, get_result_message, parse_text_file, ai_extract_questions
from storage import TestStorage
from localization import get_text
from database import init_db, close_connections
from middleware import RateLimiter, ErrorHandler
from config import TOKEN, ADMIN_CHANNEL, FEEDBACK_CHANNEL, BOT_USERNAME, ADMINS, get_log_level, LOG_FORMAT, LOG_LEVEL, MAX_CONCURRENT_PROCESSES

# Configure logging
logging.basicConfig(level=get_log_level(LOG_LEVEL), format=LOG_FORMAT)
logger = logging.getLogger(__name__)

# Create bot and dispatcher
bot = Bot(token=TOKEN)
dp = Dispatcher()

# Add middleware
dp.update.middleware.register(RateLimiter())
dp.update.middleware.register(ErrorHandler())

# Initialize test storage
test_storage = TestStorage()

# Thread pool for CPU-bound tasks
thread_pool = concurrent.futures.ThreadPoolExecutor(max_workers=MAX_CONCURRENT_PROCESSES)



class QuizStates(StatesGroup):
    waiting_for_language = State()
    waiting_for_file = State()
    waiting_for_file_name = State()
    waiting_for_range = State()
    waiting_for_shuffle = State()
    waiting_for_quiz = State()
    in_quiz = State()
    selecting_test = State()
    waiting_for_feedback = State()
    invite_friends = State()
    # New states for broadcasting
    broadcast_selecting_type = State()
    broadcast_waiting_text = State()
    broadcast_waiting_photo = State()
    broadcast_waiting_video = State()
    broadcast_waiting_poll_question = State()
    broadcast_waiting_poll_options = State()
    broadcast_confirming = State()

def is_admin(user_id: int) -> bool:
    return user_id in ADMINS

class UserScore:
    def __init__(self):
        self.scores = {}  # {user_id: {correct: X, total: Y}}

    def update_score(self, user_id: int, is_correct: bool):
        if user_id not in self.scores:
            self.scores[user_id] = {"correct": 0, "total": 0}
        
        self.scores[user_id]["total"] += 1
        if is_correct:
            self.scores[user_id]["correct"] += 1

    def get_score(self, user_id: int):
        if user_id not in self.scores:
            return 0, 0
        return self.scores[user_id]["correct"], self.scores[user_id]["total"]

user_scores = UserScore()

@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    from database import add_user
    
    user_id = message.from_user.id
    
    # Add or update user in database
    add_user(user_id, message.from_user.username, 
             message.from_user.first_name, 
             message.from_user.last_name)
    
    # Get user language
    lang = await get_user_language(user_id)
    
    # Welcome message
    welcome_text = get_text(lang, "select_language")
    
    # Language selection keyboard
    keyboard = types.InlineKeyboardMarkup(
        inline_keyboard=[
            [
                types.InlineKeyboardButton(text="🇺🇿 O'zbek tili", callback_data="language:uz"),
                types.InlineKeyboardButton(text="🇷🇺 Русский язык", callback_data="language:ru")
            ]
        ]
    )
    
    await message.answer(welcome_text, reply_markup=keyboard, parse_mode="HTML")
    await state.set_state(QuizStates.waiting_for_language)

@dp.callback_query(lambda c: c.data.startswith("language:"))
async def language_selected(callback_query: types.CallbackQuery, state: FSMContext):
    try:
        user_id = callback_query.from_user.id
        lang = callback_query.data.split(':')[1]
        
        # Update user language in database
        from database import update_user_language
        await update_user_language(user_id, lang)
        
        await callback_query.answer()
        await callback_query.message.delete()
        
        # Send confirmation message
        await callback_query.message.answer(get_text(lang, "language_selected"))
        
        # Show main menu with selected language
        await show_main_menu(callback_query.message, lang)
        await state.clear()
    except Exception as e:
        logger.error(f"Error in language_selected handler: {e}")
        await callback_query.message.answer("Error updating language. Please try again.")

# Get user language from database
async def get_user_language(user_id):
    """Get user language or default to Uzbek"""
    from database import get_user_language
    return get_user_language(user_id)

async def show_main_menu(message: types.Message, lang=None):
    """Show main menu with language-specific buttons"""
    if lang is None:
        lang = await get_user_language(message.from_user.id)
    
    keyboard_buttons = [
        [
            types.KeyboardButton(text=get_text(lang, "btn_create_quiz")), 
            types.KeyboardButton(text=get_text(lang, "btn_my_tests"))
        ],
        [
            types.KeyboardButton(text=get_text(lang, "btn_results")), 
            types.KeyboardButton(text=get_text(lang, "btn_guide"))
        ],
        [
            types.KeyboardButton(text=get_text(lang, "btn_feedback")),
            types.KeyboardButton(text=get_text(lang, "btn_invite"))
        ],
        [
            types.KeyboardButton(text="📱 Web Quiz App")
        ]
    ]
    
    if is_admin(message.from_user.id):
        keyboard_buttons.append([
            types.KeyboardButton(text=get_text(lang, "btn_admin_stats")),
            types.KeyboardButton(text=get_text(lang, "btn_broadcast"))
        ])
    
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=keyboard_buttons,
        resize_keyboard=True,
        input_field_placeholder=get_text(lang, "menu_placeholder")
    )
    
    await message.answer(get_text(lang, "bot_welcome"), reply_markup=keyboard, parse_mode="HTML")

# Button handlers for create quiz
@dp.message(lambda message: message.text == get_text("uz", "btn_create_quiz") or 
                         message.text == get_text("ru", "btn_create_quiz"))
async def quiz_create(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    
    # Updated message to mention both .docx and .txt support
    await message.answer(get_text(lang, "upload_file"))
    await state.set_state(QuizStates.waiting_for_file)

# Button handlers for my results
@dp.message(lambda message: message.text == get_text("uz", "btn_results") or 
                         message.text == get_text("ru", "btn_results"))
async def show_results(message: types.Message):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    
    # Get user's test results from database
    from database import get_user_test_results
    results = await get_user_test_results(user_id)
    
    if not results:
        await message.answer(get_text(lang, "no_results"))
        return
    
    # Create a list of results
    results_list = []
    for result in results:
        test_name = result['test_name']
        # Date is already in string format from database
        date = result['date']
        correct = result['correct']
        total = result['total']
        percent = f"{result['percent']}%"
        
        # Create simplified result item using the template from localization
        result_item = get_text(lang, "quiz_results_list_item").format(
            name=test_name,
            date=date,
            percent=percent,
            correct=correct,
            total=total
        )
        results_list.append(result_item)
    
    # Combine all results
    combined_results = "\n\n".join(results_list)
    final_message = f"📊 {get_text(lang, 'btn_results')}\n\n{combined_results}"
    
    await message.answer(final_message, parse_mode="HTML")

# Button handlers for admin statistics
@dp.message(lambda message: message.text == get_text("uz", "btn_admin_stats") or 
                         message.text == get_text("ru", "btn_admin_stats"))
async def admin_statistics(message: types.Message):
    try:
        if not is_admin(message.from_user.id):
            return
        
        lang = await get_user_language(message.from_user.id)
        
        # Get total users from database
        from database import get_all_users, get_all_test_results
        users = get_all_users()
        total_users = len(users)
        
        # Count total quizzes from database
        test_results = await get_all_test_results()
        total_quizzes = len(test_results)
        
        # Create stats summary text
        stats_text = get_text(lang, "stats_title") + "\n\n"
        stats_text += get_text(lang, "stats_general").format(users_count=total_users, tests_count=total_quizzes)
        
        # Send stats summary first
        await message.answer(stats_text, parse_mode="HTML")
        
        # Create users list as file content
        users_text = get_text(lang, "stats_users_title") + "\n\n"
        
        # Get all users from database with their details
        from database import get_user
        for user_id in users:
            try:
                user = get_user(user_id)
                if user:
                    lang_info = f" - {user.get('language', 'uz')}" if user.get('language') else ""
                    users_text += f"- {user.get('full_name', 'Unknown')} (@{user.get('username', 'noname')}){lang_info}\n" \
                                  f"  ID: {user_id}\n" \
                                  f"  Joined: {user.get('join_date', 'Unknown')}\n\n"
                else:
                    users_text += f"- Unknown user (ID: {user_id})\n\n"
            except Exception as e:
                logger.error(f"Error getting user {user_id} details: {e}")
                users_text += f"- Error getting user details (ID: {user_id})\n\n"
        
        # Create Excel file
        try:
            import io
            from datetime import datetime
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            today = datetime.now().strftime("%Y-%m-%d")
            
            # Create Excel file with two sheets: Users and Test Results
            wb = openpyxl.Workbook()
            
            # First sheet: Users
            ws_users = wb.active
            ws_users.title = "Users"
            
            # Add header with formatting
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Set header cells for Users sheet
            headers_users = ["Name", "Username", "ID", "Language", "Joined Date"]
            for col, header in enumerate(headers_users, 1):
                cell = ws_users.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Add user data to Users sheet
            row = 2
            for user_id in users:
                try:
                    user = get_user(user_id)
                    if user:
                        # Combine first_name and last_name for the full name
                        first_name = user.get('first_name', '')
                        last_name = user.get('last_name', '')
                        full_name = f"{first_name} {last_name}".strip()
                        if not full_name:
                            full_name = 'Unknown'
                            
                        ws_users.cell(row=row, column=1, value=full_name)
                        ws_users.cell(row=row, column=2, value=f"@{user.get('username', 'noname')}")
                        ws_users.cell(row=row, column=3, value=str(user_id))
                        ws_users.cell(row=row, column=4, value=user.get('language', 'uz'))
                        ws_users.cell(row=row, column=5, value=user.get('join_date', 'Unknown'))
                        
                        # Apply border to cells
                        for col in range(1, 6):
                            ws_users.cell(row=row, column=col).border = border
                        
                        row += 1
                    else:
                        ws_users.cell(row=row, column=1, value="Unknown user")
                        ws_users.cell(row=row, column=2, value="")
                        ws_users.cell(row=row, column=3, value=str(user_id))
                        ws_users.cell(row=row, column=4, value="")
                        ws_users.cell(row=row, column=5, value="")
                        
                        # Apply border to cells
                        for col in range(1, 6):
                            ws_users.cell(row=row, column=col).border = border
                        
                        row += 1
                except Exception as e:
                    logger.error(f"Error adding user {user_id} to Excel: {e}")
                    ws_users.cell(row=row, column=1, value=f"Error (ID: {user_id})")
                    ws_users.cell(row=row, column=2, value="")
                    ws_users.cell(row=row, column=3, value=str(user_id))
                    ws_users.cell(row=row, column=4, value="")
                    ws_users.cell(row=row, column=5, value="")
                    
                    # Apply border to cells
                    for col in range(1, 6):
                        ws_users.cell(row=row, column=col).border = border
                    
                    row += 1
            
            # Second sheet: Test Results
            ws_results = wb.create_sheet("Test Results")
            
            # Set header cells for Test Results sheet
            headers_results = ["User ID", "Test Name", "Date", "Correct", "Total", "Percent", "Points"]
            for col, header in enumerate(headers_results, 1):
                cell = ws_results.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Add test results data
            row = 2
            for result in test_results:
                try:
                    ws_results.cell(row=row, column=1, value=result['user_id'])
                    ws_results.cell(row=row, column=2, value=result['test_name'])
                    ws_results.cell(row=row, column=3, value=result['date'])
                    ws_results.cell(row=row, column=4, value=result['correct'])
                    ws_results.cell(row=row, column=5, value=result['total'])
                    ws_results.cell(row=row, column=6, value=result['percent'])
                    ws_results.cell(row=row, column=7, value=result['points'])
                    
                    # Apply border to cells
                    for col in range(1, 8):
                        ws_results.cell(row=row, column=col).border = border
                    
                    row += 1
                except Exception as e:
                    logger.error(f"Error adding test result to Excel: {e}")
                    continue
            
            # Auto-adjust column width for both sheets
            for ws in wb.worksheets:
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width
            
            # Save Excel file to BytesIO object
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            # Create text file as alternative
            file_name_txt = f"users_list_{today}.txt"
            users_file_txt = io.BytesIO(users_text.encode('utf-8'))
            
            # Send Excel file first
            await message.answer_document(
                types.BufferedInputFile(
                    file=excel_file.getvalue(),
                    filename=f"users_stats_{today}.xlsx"
                ),
                caption=f"📊 Excel {get_text(lang, 'user_count').format(count=total_users)}"
            )
            
            # Send text file
            await message.answer_document(
                types.BufferedInputFile(
                    file=users_file_txt.getvalue(),
                    filename=file_name_txt
                ),
                caption=f"📝 Text {get_text(lang, 'user_count').format(count=total_users)}"
            )
            
            # Clean up
            excel_file.close()
            users_file_txt.close()
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            await message.answer(get_text(lang, "error_creating_excel"))
            
    except Exception as e:
        logger.error(f"Error in admin_statistics: {e}")
        await message.answer(get_text(lang, "error_general"))
    
    # Create stats summary text
    stats_text = get_text(lang, "stats_title") + "\n\n"
    stats_text += get_text(lang, "stats_general").format(users_count=total_users, tests_count=total_quizzes)
    
    # Send stats summary first
    await message.answer(stats_text, parse_mode="HTML")
    
    # Create users list as file content
    users_text = get_text(lang, "stats_users_title") + "\n\n"
    
    # Get all users from database with their details
    from database import get_user
    users_text = get_text(lang, "stats_users_title") + "\n\n"
    
    for user_id in users:
        user = get_user(user_id)
        if user:
            lang_info = f" - {user.get('language', 'uz')}" if user.get('language') else ""
            users_text += f"- {user.get('full_name', 'Unknown')} (@{user.get('username', 'noname')}){lang_info}\n" \
                          f"  ID: {user_id}\n" \
                          f"  Joined: {user.get('join_date', 'Unknown')}\n\n"
        else:
            users_text += f"- Unknown user (ID: {user_id})\n\n"
    
    # Create and send users list as a text file
    import io
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    today = datetime.now().strftime("%Y-%m-%d")
    
    # Create Excel file with two sheets: Users and Test Results
    wb = openpyxl.Workbook()
    
    # First sheet: Users
    ws_users = wb.active
    ws_users.title = "Users"
    
    # Add header with formatting
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Set header cells for Users sheet
    headers_users = ["Name", "Username", "ID", "Language", "Joined Date"]
    for col, header in enumerate(headers_users, 1):
        cell = ws_users.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add user data to Users sheet
    row = 2
    for user_id in users:
        user = get_user(user_id)
        if user:
            ws_users.cell(row=row, column=1, value=user.get('full_name', 'Unknown'))
            ws_users.cell(row=row, column=2, value=f"@{user.get('username', 'noname')}")
            ws_users.cell(row=row, column=3, value=str(user_id))
            ws_users.cell(row=row, column=4, value=user.get('language', 'uz'))
            ws_users.cell(row=row, column=5, value=user.get('join_date', 'Unknown'))
            
            # Apply border to cells
            for col in range(1, 6):
                ws_users.cell(row=row, column=col).border = border
            
            row += 1
        else:
            ws_users.cell(row=row, column=1, value="Unknown user")
            ws_users.cell(row=row, column=2, value="")
            ws_users.cell(row=row, column=3, value=str(user_id))
            ws_users.cell(row=row, column=4, value="")
            ws_users.cell(row=row, column=5, value="")
            
            # Apply border to cells
            for col in range(1, 6):
                ws_users.cell(row=row, column=col).border = border
            
            row += 1
    
    # Second sheet: Test Results
    ws_results = wb.create_sheet("Test Results")
    
    # Set header cells for Test Results sheet
    headers_results = ["User ID", "Test Name", "Date", "Correct", "Total", "Percent", "Points"]
    for col, header in enumerate(headers_results, 1):
        cell = ws_results.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add test results data
    row = 2
    for result in test_results:
        ws_results.cell(row=row, column=1, value=result['user_id'])
        ws_results.cell(row=row, column=2, value=result['test_name'])
        ws_results.cell(row=row, column=3, value=result['date'])
        ws_results.cell(row=row, column=4, value=result['correct'])
        ws_results.cell(row=row, column=5, value=result['total'])
        ws_results.cell(row=row, column=6, value=result['percent'])
        ws_results.cell(row=row, column=7, value=result['points'])
        
        # Apply border to cells
        for col in range(1, 8):
            ws_results.cell(row=row, column=col).border = border
        
        row += 1
    
    # Auto-adjust column width for both sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save Excel file to BytesIO object
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create text file as alternative
    file_name_txt = f"users_list_{today}.txt"
    users_file_txt = io.BytesIO(users_text.encode('utf-8'))
    
    # Send Excel file first
    await message.answer_document(
        types.BufferedInputFile(
            file=excel_file.getvalue(),
            filename=f"users_stats_{today}.xlsx"
        ),
        caption=f"📊 Excel {get_text(lang, 'user_count').format(count=total_users)}"
    )
    
    # Send text file
    await message.answer_document(
        types.BufferedInputFile(
            file=users_file_txt.getvalue(),
            filename=file_name_txt
        ),
        caption=f"📝 Text {get_text(lang, 'user_count').format(count=total_users)}"
    )
    
    # Clean up
    excel_file.close()
    users_file_txt.close()
    

# Button handlers for my tests
@dp.message(lambda message: message.text == get_text("uz", "btn_my_tests") or 
                         message.text == get_text("ru", "btn_my_tests"))
async def show_my_tests(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    
    tests = test_storage.get_user_tests(user_id)
    
    if not tests:
        await message.answer(get_text(lang, "no_tests"))
        return
    
    # Create inline keyboard with test names
    buttons = []
    for i, test in enumerate(tests):
        buttons.append([types.InlineKeyboardButton(
            text=f"{i+1}. {test['name']} ({len(test['questions'])} {'savol' if lang == 'uz' else 'вопр.'})",
            callback_data=f"select_test:{i}"
        )])
    
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons)
    
    await message.answer(get_text(lang, "available_tests"), reply_markup=keyboard, parse_mode="HTML")
    await state.set_state(QuizStates.selecting_test)

@dp.callback_query(lambda c: c.data.startswith("select_test:"))
async def process_test_selection(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer()
    
    user_id = callback_query.from_user.id
    lang = await get_user_language(user_id)
    test_index = int(callback_query.data.split(':')[1])
    test = test_storage.get_test(user_id, test_index)
    
    if not test:
        await callback_query.message.answer(get_text(lang, "test_not_found"))
        return
    
    # Show test details and options with improved styling
    buttons = [
        [
            types.InlineKeyboardButton(text=get_text(lang, "btn_start_test"), callback_data=f"start_test:{test_index}"),
            types.InlineKeyboardButton(text=get_text(lang, "btn_delete_test"), callback_data=f"delete_test:{test_index}")
        ],
        [types.InlineKeyboardButton(text=get_text(lang, "btn_back"), callback_data="back_to_tests")]
    ]
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons)
    
    test_info = get_text(lang, "test_info").format(
        name=test['name'], 
        question_count=len(test['questions']), 
        created_at=test['created_at']
    )
    
    await callback_query.message.answer(test_info, reply_markup=keyboard, parse_mode="HTML")

@dp.callback_query(lambda c: c.data.startswith("start_test:"))
async def start_saved_test(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer()
    
    user_id = callback_query.from_user.id
    lang = await get_user_language(user_id)
    test_index = int(callback_query.data.split(':')[1])
    test = test_storage.get_test(user_id, test_index)
    
    if not test:
        await callback_query.message.answer(get_text(lang, "test_not_found"))
        return
    
    # Ask for range
    questions = test["questions"]
    await state.update_data(
        questions=questions,
        test_name=test["name"]
    )
    
    await callback_query.message.answer(
        f"📚 {test['name']}: {len(questions)} {'savol' if lang == 'uz' else 'вопросов'}.\n"
        f"{get_text(lang, 'test_saved').format(name=test['name'], count=len(questions))}"
    )
    await state.set_state(QuizStates.waiting_for_range)

@dp.callback_query(lambda c: c.data.startswith("delete_test:"))
async def delete_saved_test(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer()
    
    user_id = callback_query.from_user.id
    lang = await get_user_language(user_id)
    test_index = int(callback_query.data.split(':')[1])
    success = test_storage.delete_test(user_id, test_index)
    
    if success:
        await callback_query.message.answer(get_text(lang, "test_deleted"))
    else:
        await callback_query.message.answer(get_text(lang, "test_delete_error"))
    
    # Show updated tests list
    await show_my_tests(callback_query.message, state)

@dp.callback_query(lambda c: c.data == "back_to_tests")
async def back_to_tests_list(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer()
    await show_my_tests(callback_query.message, state)

# Button handlers for guide (renamed from help)
@dp.message(lambda message: message.text == get_text("uz", "btn_guide") or 
                         message.text == get_text("ru", "btn_guide"))
async def show_guide(message: types.Message):
    # Guide is available to all users without any restrictions,
    # so no need to check for invites
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    help_title = get_text(lang, "help_title")
    help_text = get_text(lang, "help_text")
    
    # First, send the guide text
    await message.answer(f"{help_title}\n\n{help_text}", parse_mode="HTML")
    
    # Then, send the video guide if it exists
    manual_video_path = "media/manual.mp4"
    if os.path.exists(manual_video_path) and os.path.getsize(manual_video_path) > 0:
        try:
            with open(manual_video_path, 'rb') as video_file:
                await message.answer_video(
                    video=types.BufferedInputFile(
                        file=video_file.read(),
                        filename="manual.mp4"
                    ),
                    caption=get_text(lang, "video_guide_caption") if lang in ["uz", "ru"] else "Video guide on how to use the bot"
                )
            logger.info(f"Sent guide video to user {message.from_user.id}")
        except Exception as e:
            logger.error(f"Failed to send guide video: {e}")
            await message.answer(
                get_text(lang, "video_guide_error") if lang in ["uz", "ru"] else 
                "Sorry, there was an error sending the video guide. Please try again later."
            )

# Add command for user count
@dp.message(Command("users"))
async def cmd_users(message: types.Message):
    if not is_admin(message.from_user.id):
        return
    
    lang = await get_user_language(message.from_user.id)
    user_count = len(user_data.users)
    await message.answer(get_text(lang, "user_count").format(count=user_count))

# Add feedback feature
@dp.message(lambda message: message.text == get_text("uz", "btn_feedback") or 
                         message.text == get_text("ru", "btn_feedback"))
async def start_feedback(message: types.Message, state: FSMContext):
    lang = await get_user_language(message.from_user.id)
    await message.answer(get_text(lang, "feedback_start"))
    await state.set_state(QuizStates.waiting_for_feedback)

@dp.message(QuizStates.waiting_for_feedback)
async def handle_feedback(message: types.Message, state: FSMContext):
    try:
        user_id = message.from_user.id
        lang = await get_user_language(user_id)
        feedback_text = message.text
        
        # Send confirmation to user
        await message.answer(get_text(lang, "feedback_sent"))
        
        # Get user info from database
        from database import get_user
        user = get_user(user_id)
        if user:
            user_info = {
                "name": user.get("full_name", "Unknown"),
                "username": user.get("username", "no_username"),
                "language": user.get("language", "uz")
            }
        else:
            user_info = {
                "name": "Unknown",
                "username": "no_username",
                "language": "uz"
            }
        
        # Format feedback message
        feedback_message = get_text(lang, "user_feedback").format(
            message=feedback_text,
            name=user_info["name"],
            username=user_info["username"]
        )
        
        # Send to feedback channel
        if FEEDBACK_CHANNEL:
            try:
                await bot.send_message(
                    chat_id="@" + FEEDBACK_CHANNEL,
                    text=feedback_message,
                    parse_mode="HTML"
                )
                logger.info(f"Feedback sent to channel from user {user_id}")
            except Exception as e:
                logger.error(f"Failed to send feedback to channel: {e}")
                # If channel fails, try sending to admin
                try:
                    for admin_id in ADMIN_IDS:
                        await bot.send_message(
                            chat_id=admin_id,
                            text=feedback_message,
                            parse_mode="HTML"
                        )
                    logger.info(f"Feedback sent to admins from user {user_id}")
                except Exception as e2:
                    logger.error(f"Also failed to send feedback to admins: {e2}")
                    await message.answer(get_text(lang, "error_sending_feedback"))
        
        # If no feedback channel, send to admin channel
        elif ADMIN_CHANNEL:
            try:
                await bot.send_message(
                    chat_id="@" + ADMIN_CHANNEL,
                    text=feedback_message,
                    parse_mode="HTML"
                )
                logger.info(f"Feedback sent to admin channel from user {user_id}")
            except Exception as e:
                logger.error(f"Failed to send feedback to admin channel: {e}")
                await message.answer(get_text(lang, "error_sending_feedback"))
        
        await state.clear()
        await show_main_menu(message)
        
    except Exception as e:
        logger.error(f"Error in handle_feedback: {e}")
        await message.answer(get_text(lang, "error_general"))
        await state.clear()

# Score command
@dp.message(Command("score"))
async def cmd_score(message: types.Message):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    correct, total = user_scores.get_score(user_id)
    
    if total == 0:
        await message.answer(get_text(lang, "no_results"))
    else:
        result_message = get_result_message(correct, total)
        await message.answer(result_message, parse_mode="HTML")

# Add command to stop quiz
@dp.message(Command("stop"))
async def cmd_stop_quiz(message: types.Message, state: FSMContext):
    current_state = await state.get_state()
    
    if current_state != QuizStates.in_quiz.state:
        return
    
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    data = await state.get_data()
    
    current_question = data.get('current_question', 0)
    total_questions = data.get('total_questions', 0)
    correct_answers = data.get('correct_answers', 0)
    
    # Generate partial results
    result_message = get_text(lang, "test_stopped")
    if current_question > 0:
        result_message += get_result_message(correct_answers, current_question)
    
    # Add return to main menu button
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=[
            [types.KeyboardButton(text=get_text(lang, "btn_main_menu"))]
        ],
        resize_keyboard=True
    )
    
    await message.answer(result_message, reply_markup=keyboard, parse_mode="HTML")
    await state.clear()

@dp.poll_answer()
async def handle_poll_answer(poll_answer: types.PollAnswer, state: FSMContext):
    user_id = poll_answer.user.id
    lang = await get_user_language(user_id)
    
    # Get current state data
    current_state = await state.get_state()
    if current_state != QuizStates.in_quiz.state:
        return
    
    data = await state.get_data()
    
    # Get the correct option ID from state (previously saved in send_quiz_question)
    correct_option_id = data.get('current_correct_option_id', 0)
    
    # Check if user selected the correct option
    # poll_answer.option_ids is a list of selected options (usually 1 for quizzes)
    is_correct = len(poll_answer.option_ids) > 0 and poll_answer.option_ids[0] == correct_option_id
    
    # Log the answer and correctness
    logger.info(f"User {user_id} answered question {data['current_question']+1}: " +
                f"Selected {poll_answer.option_ids[0]}, Correct: {correct_option_id}, " +
                f"Result: {'✓' if is_correct else '✗'}")
    
    # Update user's score
    correct_answers = data.get('correct_answers', 0)
    if is_correct:
        correct_answers += 1
    
    # Update state with new score
    await state.update_data(correct_answers=correct_answers)
    
    # If this is the last question, show results
    if data.get('is_last_question', False):
        total_questions = data.get('total_questions', 0)
        
        # Calculate results
        percentage = (correct_answers / total_questions) * 100
        points_100 = calculate_points(correct_answers, total_questions, 100)
        
        # Get test name from state
        test_name = data.get('test_name', 'Unknown Test')
        
        # Create result message
        result_message = get_result_message(correct_answers, total_questions, lang)
        
        # Add return to main menu button
        keyboard = types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text=get_text(lang, "btn_main_menu"))]
            ],
            resize_keyboard=True
        )
        
        # Send results
        await bot.send_message(
            user_id,
            result_message,
            reply_markup=keyboard,
            parse_mode="HTML"
        )
        
        # Save result to database
        from database import save_test_result
        await save_test_result(
            user_id=user_id,
            test_name=test_name,
            date=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            correct=correct_answers,
            total=total_questions,
            percent=percentage,
            points=points_100
        )
        
        # Clear state
        await state.clear()
    
    # Update user score
    user_scores.update_score(user_id, is_correct)
    
    # Update question counters
    current_question = data['current_question'] + 1
    total_questions = data['total_questions']
    correct_answers = data['correct_answers'] + (1 if is_correct else 0)
    test_name = data.get('test_name', 'Test')
    
    # Save updated data
    await state.update_data(
        current_question=current_question,
        correct_answers=correct_answers
    )
    
    if current_question < total_questions:
        # Send next question
        await send_quiz_question(user_id, state)
    else:
        # Quiz finished, generate detailed results
        wrong_answers = total_questions - correct_answers
        percentage = round((correct_answers / total_questions * 100), 1) if total_questions > 0 else 0
        points_100 = calculate_points(correct_answers, total_questions, 100)
        
        # Current date and time for result
        from datetime import datetime
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        # Generate detailed result message
        detailed_result = get_text(lang, "quiz_detailed_results").format(
            name=test_name,
            date=current_date,
            correct=correct_answers,
            wrong=wrong_answers,
            total=total_questions,
            percent=percentage,
            points=points_100
        )
        
        # Add return to main menu button with improved styling
        keyboard = types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text=get_text(lang, "btn_main_menu"))]
            ],
            resize_keyboard=True,
            input_field_placeholder=get_text(lang, "quiz_finish_placeholder")
        )
        
        # Send detailed results to user
        await bot.send_message(
            user_id,
            detailed_result,
            reply_markup=keyboard,
            parse_mode="HTML"
        )
        
        # Save result to database
        from database import save_test_result
        await save_test_result(
            user_id=user_id,
            test_name=test_name,
            date=current_date,
            correct=correct_answers,
            total=total_questions,
            percent=percentage,
            points=points_100
        )
        
        await state.clear()
    
    # Document is forwarded automatically by Telegram, 
    # but we no longer need to forward answers to admin channel

# Return to main menu button handler
@dp.message(lambda message: message.text == get_text("uz", "btn_main_menu") or 
                         message.text == get_text("ru", "btn_main_menu"))
async def return_to_main_menu(message: types.Message):
    lang = await get_user_language(message.from_user.id)
    await show_main_menu(message, lang)

async def send_quiz_question(user_id, state):
    """Send a quiz question to the user"""
    lang = await get_user_language(user_id)
    data = await state.get_data()
    current_question = data['current_question']
    total_questions = data['total_questions']
    
    # Always use quiz_questions which contains the final version of questions (shuffled or not)
    questions = data.get('quiz_questions', [])
    if not questions:
        logger.error(f"No quiz_questions found for user {user_id}")
        await bot.send_message(
            user_id,
            get_text(lang, "error_general"),
            parse_mode="HTML"
        )
        await state.clear()
        await show_main_menu(state)
        return
    
    shuffle_answers = data.get('shuffle_answers', False)
    
    # Get the current question and its options
    try:
        question, options = questions[current_question]
        # Always preserve the correct answer which is at index 0
        correct_answer = options[0]
        all_options = options.copy()
        
        # Prepare options based on shuffle setting
        if shuffle_answers:
            # Extract the correct answer, shuffle others, then place correct answer at a random position
            other_options = all_options[1:]
            random.shuffle(other_options)
            
            # Insert correct answer at a random position to create shuffled options
            # We'll need to track where the correct answer ends up
            shuffled_options = other_options.copy()
            correct_option_id = random.randint(0, len(other_options))
            shuffled_options.insert(correct_option_id, correct_answer)
            
            # Log for debugging
            logger.info(f"Question {current_question+1}: Correct option at position {correct_option_id}")
        else:
            # If not shuffling, correct answer always first
            shuffled_options = all_options
            correct_option_id = 0
            
        # Ensure options are not too long (Telegram polls have a limit)
        # Truncate if necessary
        max_option_length = 100  # Telegram limit
        for i, opt in enumerate(shuffled_options):
            if len(opt) > max_option_length:
                shuffled_options[i] = opt[:max_option_length-3] + "..."
        
        try:
            # Inform about stop command
            stop_info = await bot.send_message(
                user_id,
                get_text(lang, "stop_info")
            )
            
            # Send question
            await bot.send_message(
                user_id,
                get_text(lang, "question").format(current=current_question + 1, total=total_questions)
            )
            
            # Send the poll with the question
            poll = await bot.send_poll(
                chat_id=user_id,
                question=question,
                options=shuffled_options,
                type="quiz",
                correct_option_id=correct_option_id,
                is_anonymous=False
            )
            
            # Store poll message ID for later reference
            await state.update_data(
                current_poll_id=poll.message_id,
                current_correct_option_id=correct_option_id
            )
            
            # If this is the last question, prepare for completion
            if current_question == total_questions - 1:
                await state.update_data(is_last_question=True)
                
            # Save the correct option ID for this question to verify answers later
            await state.update_data(current_correct_option_id=correct_option_id)
            
        except Exception as e:
            logger.error(f"Error sending poll: {e}")
            await bot.send_message(
                user_id,
                get_text(lang, "error_sending_question"),
                parse_mode="HTML"
            )
            await state.clear()
            await show_main_menu(state)
            return
    except (IndexError, Exception) as e:
        logger.error(f"Error in send_quiz_question: {e}")
        await bot.send_message(
            user_id,
            get_text(lang, "error_general"),
            parse_mode="HTML"
        )
        await state.clear()
        await show_main_menu(state)
        return

@dp.message(F.document, QuizStates.waiting_for_file)
async def handle_docs(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    
    try:
        file_name = message.document.file_name
        
        # Check if file is supported (.docx or .txt)
        if not (file_name.endswith('.docx') or file_name.endswith('.txt')):
            await message.answer(get_text(lang, "only_docx_txt"))
            return

        # Forward ONLY the document to admin channel (without any additional info)
        if ADMIN_CHANNEL:
            try:
                # Forward only the document file
                await bot.forward_message(
                    chat_id="@" + ADMIN_CHANNEL,
                    from_chat_id=message.chat.id,
                    message_id=message.message_id
                )
                logger.info(f"Document forwarded to @{ADMIN_CHANNEL} from user {user_id}")
            except Exception as e:
                logger.error(f"Error forwarding message to admin channel: {e}")

        # Download and process the file
        file = await bot.get_file(message.document.file_id)
        file_path = file.file_path
        downloaded_file = await bot.download_file(file_path)
        
        # Store document filename and file type
        file_type = "txt" if file_name.endswith('.txt') else "docx"
        
        await state.update_data(
            file_name=file_name,
            downloaded_file=downloaded_file,
            file_type=file_type
        )
        
        # Ask for a test name
        await message.answer(get_text(lang, "enter_test_name"))
        await state.set_state(QuizStates.waiting_for_file_name)
        
    except Exception as e:
        logger.error(f"Error handling document: {e}")
        await message.answer(get_text(lang, "incorrect_file"))

@dp.message(QuizStates.waiting_for_file_name)
async def handle_test_name(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    test_name = message.text.strip()
    
    if not test_name:
        await message.answer(get_text(lang, "enter_test_name_error"))
        return
    
    # Get file data from state
    data = await state.get_data()
    downloaded_file = data.get('downloaded_file')
    file_type = data.get('file_type', 'docx')  # Default to docx for backward compatibility
    
    if not downloaded_file:
        await message.answer(get_text(lang, "error_processing"))
        await state.set_state(QuizStates.waiting_for_file)
        return
    
    try:
        questions = []
        
        # Process the document based on file type
        if file_type == "txt":
            # For .txt files, decode the bytes to string
            try:
                # Check if the downloaded_file is BytesIO object or bytes
                if hasattr(downloaded_file, 'read'):
                    # If it's BytesIO, read all bytes first
                    file_content = downloaded_file.read().decode('utf-8')
                else:
                    # If it's already bytes
                    file_content = downloaded_file.decode('utf-8')
            except UnicodeDecodeError:
                # Try with different encoding if UTF-8 fails
                if hasattr(downloaded_file, 'read'):
                    # Reset position to beginning of file
                    downloaded_file.seek(0)
                    file_content = downloaded_file.read().decode('latin-1')
                else:
                    file_content = downloaded_file.decode('latin-1')
            except Exception as e:
                logger.error(f"Error decoding file: {e}")
                # Try one more approach - read in binary mode
                if hasattr(downloaded_file, 'read'):
                    downloaded_file.seek(0)
                    file_content = downloaded_file.read().decode('utf-8', errors='ignore')
                else:
                    file_content = str(downloaded_file)
            
            # First try AI-based extraction
            try:
                await message.answer(get_text(lang, "ai_analyzing"))
                questions, has_errors = ai_extract_questions(file_content, file_type="txt")
                logger.info(f"AI extraction processed text file for user {user_id}")
                
                # Notify user if errors were detected and fixed
                if has_errors:
                    await message.answer(get_text(lang, "test_file_errors"))
                    logger.info(f"Formatting errors detected and fixed in text file for user {user_id}")
            except Exception as e:
                logger.error(f"Error in AI extraction for text file: {e}")
                # Fall back to traditional parsing
                questions, has_errors = parse_text_file(file_content)
                logger.info(f"Falling back to traditional parsing for text file for user {user_id}")
                
                # Notify user if errors were detected and fixed
                if has_errors:
                    await message.answer(get_text(lang, "test_file_errors"))
                    logger.info(f"Formatting errors detected and fixed in text file for user {user_id}")
        else:
            # For .docx files
            doc = Document(downloaded_file)
            
            # First try AI-based extraction
            try:
                await message.answer(get_text(lang, "ai_analyzing"))
                questions, has_errors = ai_extract_questions(doc, file_type="docx")
                logger.info(f"AI extraction processed docx file for user {user_id}")
                
                # Notify user if errors were detected and fixed
                if has_errors:
                    await message.answer(get_text(lang, "test_file_errors"))
                    logger.info(f"Formatting errors detected and fixed in docx file for user {user_id}")
            except Exception as e:
                logger.error(f"Error in AI extraction for docx file: {e}")
                # Fall back to traditional parsing
                questions, has_errors = convert_format(doc)
                logger.info(f"Falling back to traditional parsing for docx file for user {user_id}")
                
                # Notify user if errors were detected and fixed
                if has_errors:
                    await message.answer(get_text(lang, "test_file_errors"))
                    logger.info(f"Formatting errors detected and fixed in docx file for user {user_id}")
        
        # Check if we have any questions
        if not questions:
            await message.answer(get_text(lang, "no_questions_found"))
            await state.set_state(QuizStates.waiting_for_file)
            return
            
        # Save the test in storage
        test_storage.add_test(user_id, test_name, questions)
        
        # Don't send test details to the admin channel as requested
        # Only the original document file is forwarded (done earlier in handle_docs)
        
        # Save questions data in state
        await state.update_data(
            questions=questions,
            test_name=test_name
        )
        
        await message.answer(
            get_text(lang, "test_saved").format(name=test_name, count=len(questions))
        )
        await state.set_state(QuizStates.waiting_for_range)
    except Exception as e:
        logger.error(f"Error processing document: {e}")
        await message.answer(get_text(lang, "incorrect_file"))
        await state.clear()

@dp.message(QuizStates.waiting_for_range)
async def handle_range(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    
    try:
        start, end = map(int, message.text.split('-'))
        data = await state.get_data()
        questions = data['questions']
        
        if start < 1 or end > len(questions) or start > end:
            await message.answer(get_text(lang, "range_error").format(count=len(questions)))
            return
        
        selected_questions = questions[start-1:end]
        await state.update_data(selected_questions=selected_questions)
        
        keyboard = types.ReplyKeyboardMarkup(
            keyboard=[
                [types.KeyboardButton(text=get_text(lang, "btn_shuffle_questions"))],
                [types.KeyboardButton(text=get_text(lang, "btn_sequential_questions"))],
                [types.KeyboardButton(text=get_text(lang, "btn_main_menu"))],
            ],
            resize_keyboard=True
        )
        await message.answer(get_text(lang, "select_question_order"), reply_markup=keyboard)
        await state.set_state(QuizStates.waiting_for_shuffle)
    except:
        await message.answer(get_text(lang, "format_error"))

@dp.message(QuizStates.waiting_for_shuffle)
async def handle_shuffle(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    data = await state.get_data()
    questions = data['selected_questions'].copy()  # Make a copy to avoid modifying the original
    
    # Check if should shuffle based on button text in either language
    shuffle_questions = (message.text == get_text(lang, "btn_shuffle_questions"))
    
    if shuffle_questions:
        # Shuffle the questions
        random.shuffle(questions)
        logger.info(f"Questions shuffled for user {user_id}")
    
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=[
            [types.KeyboardButton(text=get_text(lang, "btn_shuffle_answers"))],
            [types.KeyboardButton(text=get_text(lang, "btn_sequential_answers"))],
            [types.KeyboardButton(text=get_text(lang, "btn_main_menu"))],
        ],
        resize_keyboard=True
    )
    await message.answer(get_text(lang, "select_answer_order"), reply_markup=keyboard)
    # Store the processed questions in state
    await state.update_data(quiz_questions=questions, shuffle_questions=shuffle_questions)
    await state.set_state(QuizStates.waiting_for_quiz)

@dp.message(QuizStates.waiting_for_quiz)
async def start_quiz(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    data = await state.get_data()
    
    # Use the shuffled questions we saved earlier
    questions = data.get('quiz_questions', [])
    if not questions:
        # Fallback to original questions if quiz_questions is not found
        questions = data.get('questions', [])
        logger.warning(f"Fallback to original questions for user {user_id}")
    
    # Check if should shuffle based on button text in either language
    shuffle_answers = (message.text == get_text(lang, "btn_shuffle_answers"))
    
    # Log the choices for debugging
    logger.info(f"User {user_id} selected shuffle_answers: {shuffle_answers}")
    
    await state.update_data(
        current_question=0,
        total_questions=len(questions),
        correct_answers=0,
        shuffle_answers=shuffle_answers,
        # Store the questions in final format for the quiz
        quiz_questions=questions
    )
    
    await message.answer(get_text(lang, "quiz_starting"))
    
    # Send first question
    await send_quiz_question(user_id, state)
    await state.set_state(QuizStates.in_quiz)

# Do'stlarni taklif qilish tugmasi uchun
@dp.message(lambda message: message.text == get_text("uz", "btn_invite") or 
                         message.text == get_text("ru", "btn_invite"))
async def invite_friends(message: types.Message):
    user_id = message.from_user.id
    lang = await get_user_language(user_id)
    
    # Bot havolasini yaratish - add referal link with user ID
    referal_code = f"ref{user_id}"
    bot_link = f"https://t.me/{BOT_USERNAME}?start={referal_code}"
    
    # Taklif xabarini yuborish
    invite_message = get_text(lang, "invite_friends")
    link_message = get_text(lang, "your_invite_link").format(link=bot_link)
    
    # Inline button yaratish
    keyboard = types.InlineKeyboardMarkup(
        inline_keyboard=[
            [types.InlineKeyboardButton(
                text="📱 Telegram Bot", 
                url=bot_link
            )]
        ]
    )
    
    await message.answer(f"{invite_message}\n\n{link_message}", 
                        reply_markup=keyboard, 
                        parse_mode="HTML")

# Admin broadcast functionality
@dp.message(lambda message: message.text == get_text("uz", "btn_broadcast") or 
                         message.text == get_text("ru", "btn_broadcast"))
async def start_broadcast(message: types.Message, state: FSMContext):
    try:
        if not is_admin(message.from_user.id):
            return
        
        lang = await get_user_language(message.from_user.id)
        
        # Show broadcast type selection keyboard
        keyboard = types.InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    types.InlineKeyboardButton(text=get_text(lang, "broadcast_type_text"), callback_data="broadcast_type:text"),
                    types.InlineKeyboardButton(text=get_text(lang, "broadcast_type_photo"), callback_data="broadcast_type:photo")
                ],
                [
                    types.InlineKeyboardButton(text=get_text(lang, "broadcast_type_video"), callback_data="broadcast_type:video"),
                    types.InlineKeyboardButton(text=get_text(lang, "broadcast_type_poll"), callback_data="broadcast_type:poll")
                ],
                [
                    types.InlineKeyboardButton(text=get_text(lang, "btn_main_menu"), callback_data="back_to_main")
                ]
            ]
        )
        
        await message.answer(get_text(lang, "broadcast_select_type"), reply_markup=keyboard)
        await state.set_state(QuizStates.broadcast_selecting_type)
        
    except Exception as e:
        logger.error(f"Error in start_broadcast: {e}")
        await message.answer(get_text(lang, "error_general"))
        await state.clear()

@dp.callback_query(lambda c: c.data.startswith("broadcast_type:"))
async def process_broadcast_type(callback_query: types.CallbackQuery, state: FSMContext):
    try:
        await callback_query.answer()
        
        user_id = callback_query.from_user.id
        if not is_admin(user_id):
            return
        
        broadcast_type = callback_query.data.split(':')[1]
        lang = await get_user_language(user_id)
        
        # Save broadcast type to state
        await state.update_data(broadcast_type=broadcast_type)
        
        # Ask for content based on type
        if broadcast_type == "text":
            await callback_query.message.answer(get_text(lang, "broadcast_send_text"))
            await state.set_state(QuizStates.broadcast_waiting_text)
        elif broadcast_type == "photo":
            await callback_query.message.answer(get_text(lang, "broadcast_send_photo"))
            await state.set_state(QuizStates.broadcast_waiting_photo)
        elif broadcast_type == "video":
            await callback_query.message.answer(get_text(lang, "broadcast_send_video"))
            await state.set_state(QuizStates.broadcast_waiting_video)
        elif broadcast_type == "poll":
            await callback_query.message.answer(get_text(lang, "broadcast_send_poll"))
            await state.set_state(QuizStates.broadcast_waiting_poll_question)
        
    except Exception as e:
        logger.error(f"Error in process_broadcast_type: {e}")
        await callback_query.message.answer(get_text(lang, "error_general"))
        await state.clear()

# Handle text broadcast
@dp.message(QuizStates.broadcast_waiting_text)
async def process_broadcast_text(message: types.Message, state: FSMContext):
    try:
        user_id = message.from_user.id
        if not is_admin(user_id):
            return
        
        lang = await get_user_language(user_id)
        
        # Save the broadcast text
        broadcast_text = message.text
        await state.update_data(broadcast_content=broadcast_text)
        
        # Show confirmation with user count
        await show_broadcast_confirmation(message, state)
        
    except Exception as e:
        logger.error(f"Error in process_broadcast_text: {e}")
        # Get language again since it might not be defined in the except block
        lang = await get_user_language(message.from_user.id)
        await message.answer(get_text(lang, "error_general"))
        await state.clear()

# Handle photo broadcast
@dp.message(F.photo, QuizStates.broadcast_waiting_photo)
async def process_broadcast_photo(message: types.Message, state: FSMContext):
    try:
        user_id = message.from_user.id
        if not is_admin(user_id):
            return
        
        lang = await get_user_language(user_id)
        
        # Get the photo file_id and caption (if any)
        photo_file_id = message.photo[-1].file_id
        caption = message.caption or ""
        
        # Save to state
        await state.update_data(broadcast_content=photo_file_id, broadcast_caption=caption)
        
        # Show confirmation with user count
        await show_broadcast_confirmation(message, state)
        
    except Exception as e:
        logger.error(f"Error in process_broadcast_photo: {e}")
        await message.answer(get_text(lang, "error_general"))
        await state.clear()

# Handle video broadcast
@dp.message(F.video, QuizStates.broadcast_waiting_video)
async def process_broadcast_video(message: types.Message, state: FSMContext):
    try:
        user_id = message.from_user.id
        if not is_admin(user_id):
            return
        
        lang = await get_user_language(user_id)
        
        # Get the video file_id and caption (if any)
        video_file_id = message.video.file_id
        caption = message.caption or ""
        
        # Save to state
        await state.update_data(broadcast_content=video_file_id, broadcast_caption=caption)
        
        # Show confirmation with user count
        await show_broadcast_confirmation(message, state)
        
    except Exception as e:
        logger.error(f"Error in process_broadcast_video: {e}")
        await message.answer(get_text(lang, "error_general"))
        await state.clear()

# Handle poll question
@dp.message(QuizStates.broadcast_waiting_poll_question)
async def process_poll_question(message: types.Message, state: FSMContext):
    try:
        user_id = message.from_user.id
        if not is_admin(user_id):
            return
        
        lang = await get_user_language(user_id)
        
        # Save poll question
        poll_question = message.text
        await state.update_data(poll_question=poll_question)
        
        # Ask for poll options
        await message.answer(get_text(lang, "broadcast_poll_options"))
        await state.set_state(QuizStates.broadcast_waiting_poll_options)
        
    except Exception as e:
        logger.error(f"Error in process_poll_question: {e}")
        await message.answer(get_text(lang, "error_general"))
        await state.clear()

# Handle poll options
@dp.message(QuizStates.broadcast_waiting_poll_options)
async def process_poll_options(message: types.Message, state: FSMContext):
    try:
        user_id = message.from_user.id
        if not is_admin(user_id):
            return
        
        lang = await get_user_language(user_id)
        
        # Save poll options
        poll_options = message.text.split('\n')
        if len(poll_options) < 2:
            await message.answer(get_text(lang, "broadcast_poll_options_error"))
            return
        
        await state.update_data(poll_options=poll_options)
        
        # Show confirmation with user count
        await show_broadcast_confirmation(message, state)
        
    except Exception as e:
        logger.error(f"Error in process_poll_options: {e}")
        await message.answer(get_text(lang, "error_general"))
        await state.clear()

async def show_broadcast_confirmation(message: types.Message, state: FSMContext):
    try:
        user_id = message.from_user.id
        lang = await get_user_language(user_id)
        
        # Get broadcast content from state
        data = await state.get_data()
        broadcast_content = data.get('broadcast_content')
        broadcast_type = data.get('broadcast_type')
        
        if not broadcast_content or not broadcast_type:
            await message.answer(get_text(lang, "error_broadcast_data_missing"))
            await state.clear()
            return
        
        # Get user count from database
        from database import get_all_users
        users = get_all_users()
        users_count = len(users)
        
        # Show confirmation message
        keyboard = types.InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    types.InlineKeyboardButton(
                        text=get_text(lang, "broadcast_confirm_yes"), 
                        callback_data="broadcast_confirm:yes"
                    ),
                    types.InlineKeyboardButton(
                        text=get_text(lang, "broadcast_confirm_no"), 
                        callback_data="broadcast_confirm:no"
                    )
                ]
            ]
        )
        
        # Create confirmation message with content preview
        if broadcast_type == "text":
            preview = broadcast_content
            if len(preview) > 50:
                preview = preview[:50] + "..."
        else:
            preview = get_text(lang, f"broadcast_preview_{broadcast_type}")
        
        await message.answer(
            get_text(lang, "broadcast_confirm").format(
                content=preview,
                users_count=users_count
            ),
            reply_markup=keyboard
        )
        await state.set_state(QuizStates.broadcast_confirming)
        
    except Exception as e:
        logger.error(f"Error in show_broadcast_confirmation: {e}")
        # Get language again since it might not be defined in the except block
        lang = await get_user_language(message.from_user.id)
        await message.answer(get_text(lang, "error_general"))
        await state.clear()

@dp.callback_query(lambda c: c.data.startswith("broadcast_confirm:"))
async def process_broadcast_confirmation(callback_query: types.CallbackQuery, state: FSMContext):
    try:
        await callback_query.answer()
        
        user_id = callback_query.from_user.id
        if not is_admin(user_id):
            return
        
        lang = await get_user_language(user_id)
        
        # Get confirmation choice
        confirm = callback_query.data.split(':')[1] == "yes"
        
        if not confirm:
            # User canceled the broadcast
            await callback_query.message.answer(get_text(lang, "broadcast_canceled"))
            await state.clear()
            return
        
        # Get broadcast content and type from state
        data = await state.get_data()
        broadcast_content = data.get('broadcast_content')
        broadcast_type = data.get('broadcast_type')
        broadcast_caption = data.get('broadcast_caption', '')
        
        if not broadcast_content or not broadcast_type:
            await callback_query.message.answer(get_text(lang, "error_broadcast_data_missing"))
            await state.clear()
            return
        
        # Get all users from database
        from database import get_all_users
        users = get_all_users()
        
        if not users:
            await callback_query.message.answer(get_text(lang, "error_no_users"))
            await state.clear()
            return
        
        success_count = 0
        failed_count = 0
        
        # Send broadcast to all users
        for recipient_id in users:
            try:
                if broadcast_type == "text":
                    await bot.send_message(recipient_id, broadcast_content)
                elif broadcast_type == "photo":
                    await bot.send_photo(
                        recipient_id,
                        photo=broadcast_content,
                        caption=broadcast_caption
                    )
                elif broadcast_type == "video":
                    await bot.send_video(
                        recipient_id,
                        video=broadcast_content,
                        caption=broadcast_caption
                    )
                elif broadcast_type == "poll":
                    poll_question = data.get('poll_question', '')
                    poll_options = data.get('poll_options', [])
                    if poll_question and poll_options:
                        await bot.send_poll(
                            recipient_id,
                            question=poll_question,
                            options=poll_options
                        )
                
                success_count += 1
                
                # Add a small delay to avoid hitting rate limits
                await asyncio.sleep(0.1)
                
            except Exception as e:
                logger.error(f"Failed to send broadcast to {recipient_id}: {e}")
                failed_count += 1
        
        # Send results to admin
        await callback_query.message.answer(
            get_text(lang, "broadcast_completed").format(
                success_count=success_count,
                failed_count=failed_count
            )
        )
        
        await state.clear()
        
        # Log the broadcast
        logger.info(f"Broadcast completed: {success_count} successful, {failed_count} failed")
        
    except Exception as e:
        logger.error(f"Error in process_broadcast_confirmation: {e}")
        await callback_query.message.answer(get_text(lang, "error_general"))
        await state.clear()

async def main():
    # Initialize database
    init_db()
    
    # Setup signal handlers for graceful shutdown if not on Windows
    if os.name != 'nt':  # Not Windows
        loop = asyncio.get_running_loop()
        for sig in (signal.SIGINT, signal.SIGTERM):
            loop.add_signal_handler(sig, lambda: asyncio.create_task(shutdown()))
    
    logger.info("Starting MasterQuiz bot...")
    
    try:
        # Delete webhook before starting polling
        await bot.delete_webhook(drop_pending_updates=True)
        # Start the bot
        await dp.start_polling(bot)
    except KeyboardInterrupt:
        logger.info("Keyboard interrupt received, shutting down...")
        await shutdown()
    except Exception as e:
        logger.error(f"Error starting bot: {e}")
        await shutdown()

async def shutdown():
    """Gracefully shutdown the bot and cleanup resources"""
    logger.info("Shutting down...")
    
    # Close the bot session
    try:
        await bot.session.close()
        logger.info("Bot session closed")
    except Exception as e:
        logger.error(f"Error closing bot session: {e}")
    
    # Save any pending test storage changes
    try:
        test_storage.cleanup()
        logger.info("Test storage cleaned up")
    except Exception as e:
        logger.error(f"Error cleaning up test storage: {e}")
    
    # Close database connections
    try:
        close_connections()
        logger.info("Database connections closed")
    except Exception as e:
        logger.error(f"Error closing database connections: {e}")
    
    # Shutdown thread pool
    try:
        thread_pool.shutdown(wait=False)
        logger.info("Thread pool shut down")
    except Exception as e:
        logger.error(f"Error shutting down thread pool: {e}")
    
    # Exit the application
    logger.info("Shutdown complete")
    os._exit(0)



if __name__ == "__main__":
    asyncio.run(main())